from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).resolve().parents[1]
CSV_PATH = BASE / 'data' / 'products.csv'
XLSX_PATH = BASE / 'data' / 'products_master.xlsx'

with CSV_PATH.open('r', encoding='utf-8', newline='') as f:
    rows = list(csv.DictReader(f, delimiter='|'))

wb = Workbook()
ws = wb.active
ws.title = 'Product_Master'
headers = ['Family', 'Subcategory', 'Product', 'Applications', 'Typical Spec', 'Grades / Forms', 'Image File']
ws.append(headers)
for row in rows:
    ws.append([row.get(h, '') for h in headers])

for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='0A2540')
    cell.alignment = Alignment(horizontal='center')

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Image repository sheet
img_ws = wb.create_sheet('Image_References')
image_files = sorted([p.name for p in (BASE / 'assets' / 'images').glob('*.svg')])
img_ws.append(['Image File', 'Repository Folder', 'Suggested Product Association'])
for image in image_files:
    img_ws.append([image, 'assets/images', 'Map in Product_Master'])

for cell in img_ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='C8102E')

for row in img_ws.iter_rows(min_row=2, max_row=img_ws.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Auto-fit widths
for sheet in wb.worksheets:
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 40)

wb.save(XLSX_PATH)
print(f'Created workbook: {XLSX_PATH}')
