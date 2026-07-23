from pathlib import Path
import csv
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[1]
XLSX_PATH = BASE / 'data' / 'products_master.xlsx'
CSV_PATH = BASE / 'data' / 'products.csv'

wb = load_workbook(XLSX_PATH, data_only=True)
ws = wb['Product_Master']
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
final_rows = []
for row in rows[1:]:
    if not any(cell is not None and str(cell).strip() for cell in row):
        continue
    entry = dict(zip(headers, row))
    final_rows.append(entry)

with CSV_PATH.open('w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=headers, delimiter='|')
    writer.writeheader()
    writer.writerows(final_rows)

print(f'Synced {len(final_rows)} product rows from Excel to CSV.')
